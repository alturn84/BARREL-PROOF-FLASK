#!/usr/bin/env python3
"""
MLB Batting Stats Updater — Barrel Proof Vault
───────────────────────────────────────────────
Pulls current-season batting stats from Baseball Reference via pybaseball
and writes/overwrites a formatted Excel file.

Setup (one time):
    pip install pybaseball openpyxl pandas

Usage:
    python update_batting_stats.py              # current season
    python update_batting_stats.py 2025         # specific season

Schedule with cron (runs at 9 AM daily):
    0 9 * * * /Library/Frameworks/Python.framework/Versions/3.14/bin/python3 "/Users/allanturner/Documents/BARREL PROOF/update_batting_stats.py"
"""

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pybaseball import batting_stats_bref
from pybaseball import cache

# ── Config ────────────────────────────────────────────────────────────────────
VAULT   = Path(".")
OUTPUT  = VAULT / "2026 Batting Stats" / "mlb_batting_stats.xlsx"
MIN_PA  = 50

cache.enable()
cache.purge()

def fix_name(val):
    """Fix pybaseball literal \\xNN escape sequences and strip B-Ref symbols."""
    if not isinstance(val, str):
        return val
    try:
        # pybaseball returns e.g. 'Acu\\xc3\\xb1a' — decode escape sequences then re-decode as UTF-8
        val = bytes(val, 'utf-8').decode('unicode_escape').encode('latin-1').decode('utf-8')
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    val = unicodedata.normalize('NFC', val)
    val = val.replace('*', '').replace('#', '').strip()
    return val

# ── Fetch ─────────────────────────────────────────────────────────────────────
season = int(sys.argv[1]) if len(sys.argv) > 1 else datetime.today().year
print(f"Fetching {season} batting stats from Baseball Reference...")

df = batting_stats_bref(season)

keep = ['Name','Age','Tm','G','PA','AB','R','H','2B','3B','HR','RBI',
        'BB','IBB','SO','HBP','SH','SF','GDP','SB','CS','BA','OBP','SLG','OPS']

df = df[[c for c in keep if c in df.columns]].copy()
df = df.rename(columns={'Tm': 'Team', 'GDP': 'GIDP'})
df['Name'] = df['Name'].apply(fix_name)

df = df[df['PA'] >= MIN_PA].sort_values('PA', ascending=False).reset_index(drop=True)
df.insert(0, 'Rk', range(1, len(df) + 1))

print(f"  {len(df)} players with {MIN_PA}+ PA found.")
accented = [n for n in df['Name'] if any(ord(c) > 127 for c in n)]
print(f"  Sample accented names: {accented[:5]}")

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"MLB Batting {season}"

header_fill  = PatternFill('solid', start_color='1F4E79')
header_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
alt_fill     = PatternFill('solid', start_color='DDEEFF')
white_fill   = PatternFill('solid', start_color='FFFFFF')
body_font    = Font(name='Calibri', size=10)
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin         = Side(style='thin', color='AAAAAA')
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = list(df.columns)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

ws.row_dimensions[1].height = 28

for row_idx, (_, row) in enumerate(df.iterrows(), 2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    ws.row_dimensions[row_idx].height = 20
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val  = row[col_name]
        if hasattr(val, 'item'):
            val = val.item()
        cell.value     = val
        cell.font      = body_font
        cell.fill      = fill
        cell.border    = border
        cell.alignment = left_align if col_name == 'Name' else center_align
        if isinstance(val, float):
            fmt = {
                'BA':  '0.000',
                'OBP': '0.000',
                'SLG': '0.000',
                'OPS': '0.000',
            }.get(col_name)
            if fmt:
                cell.number_format = fmt

widths = {
    'Rk':5,'Name':22,'Age':6,'Team':15,'G':5,'PA':5,'AB':5,
    'R':5,'H':5,'2B':5,'3B':5,'HR':5,'RBI':5,'BB':5,'IBB':5,
    'SO':5,'HBP':5,'SH':5,'SF':5,'GIDP':6,'SB':5,'CS':5,
    'BA':8,'OBP':8,'SLG':8,'OPS':8
}
for col_idx, col_name in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 7)

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

meta = wb.create_sheet("Info")
meta['A1'] = "Last updated"
meta['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M")
meta['A2'] = "Season"
meta['B2'] = season
meta['A3'] = "Min PA filter"
meta['B3'] = MIN_PA
meta['A4'] = "Source"
meta['B4'] = "Baseball Reference via pybaseball"
for c in ['A1','A2','A3','A4']:
    meta[c].font = Font(bold=True, name='Calibri')

wb.save(OUTPUT)
print(f"\n✓  Saved → {OUTPUT}")
print(f"   Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
