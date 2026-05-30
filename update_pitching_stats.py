#!/usr/bin/env python3
"""
MLB Pitching Stats Updater — Barrel Proof Vault
────────────────────────────────────────────────
Pulls current-season pitching stats from Baseball Reference via pybaseball
and writes/overwrites a formatted Excel file.

Setup (one time):
    pip install pybaseball openpyxl pandas

Usage:
    python update_pitching_stats.py              # current season
    python update_pitching_stats.py 2025         # specific season

Schedule with cron (runs at 9:05 AM daily):
    5 9 * * * /Library/Frameworks/Python.framework/Versions/3.14/bin/python3 "/Users/allanturner/Documents/BARREL PROOF/update_pitching_stats.py"
"""

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pybaseball import pitching_stats_bref
from pybaseball import cache

# ── Config ────────────────────────────────────────────────────────────────────
VAULT   = Path("/Users/allanturner/BARREL PROOF")
OUTPUT  = VAULT / "2026 Pitching Stats" / "mlb_pitching_stats.xlsx"
MIN_IP  = 10

cache.enable()
cache.purge()

def fix_name(val):
    """Fix pybaseball literal \\xNN escape sequences and strip B-Ref symbols."""
    if not isinstance(val, str):
        return val
    try:
        val = bytes(val, 'utf-8').decode('unicode_escape').encode('latin-1').decode('utf-8')
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    val = unicodedata.normalize('NFC', val)
    val = val.replace('*', '').replace('#', '').strip()
    return val

# ── Fetch ─────────────────────────────────────────────────────────────────────
season = int(sys.argv[1]) if len(sys.argv) > 1 else datetime.today().year
print(f"Fetching {season} pitching stats from Baseball Reference...")

df = pitching_stats_bref(season)

keep = ['Name','Age','Tm','W','L','G','GS','SV','IP','H','R','ER',
        'HR','BB','IBB','SO','HBP','BF','ERA','WHIP','SO9','SO/W',
        'BAbip','GB/FB']

df = df[[c for c in keep if c in df.columns]].copy()
df = df.rename(columns={
    'Tm':    'Team',
    'SO/W':  'SO/BB',
    'BAbip': 'BABIP',
})

df['Name'] = df['Name'].apply(fix_name)

df = df[df['IP'] >= MIN_IP].sort_values('IP', ascending=False).reset_index(drop=True)
df.insert(0, 'Rk', range(1, len(df) + 1))

print(f"  {len(df)} pitchers with {MIN_IP}+ IP found.")
accented = [n for n in df['Name'] if any(ord(c) > 127 for c in n)]
print(f"  Sample accented names: {accented[:5]}")

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"MLB Pitching {season}"

header_fill  = PatternFill('solid', start_color='1A3A2A')
header_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
alt_fill     = PatternFill('solid', start_color='D6EAD6')
white_fill   = PatternFill('solid', start_color='FFFFFF')
body_font    = Font(name='Calibri', size=10)
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin         = Side(style='thin', color='AAAAAA')
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

headers   = list(df.columns)
rate_cols = {'ERA','WHIP','SO9','SO/BB','BABIP','GB/FB'}

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

ws.row_dimensions[1].height = 22

for row_idx, (_, row) in enumerate(df.iterrows(), 2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val  = row[col_name]
        if hasattr(val, 'item'):
            val = val.item()
        cell.value     = val
        cell.font      = body_font
        cell.fill      = fill
        cell.border    = border
        cell.alignment = left_align if col_name == 'Name' else center_align
        if col_name in rate_cols and isinstance(val, float):
            cell.number_format = '0.00'

col_widths = {
    'Rk':5,'Name':24,'Age':5,'Team':16,'W':4,'L':4,'G':5,'GS':5,
    'SV':5,'IP':6,'H':5,'R':5,'ER':5,'HR':5,'BB':5,'IBB':5,
    'SO':5,'HBP':5,'BF':5,'ERA':6,'WHIP':6,'SO9':6,'SO/BB':7,
    'BABIP':7,'GB/FB':7
}
for col_idx, col_name in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 7)

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

meta = wb.create_sheet("Info")
meta['A1'] = "Last updated"
meta['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M")
meta['A2'] = "Season"
meta['B2'] = season
meta['A3'] = "Min IP filter"
meta['B3'] = MIN_IP
meta['A4'] = "Source"
meta['B4'] = "Baseball Reference via pybaseball"
for c in ['A1','A2','A3','A4']:
    meta[c].font = Font(bold=True, name='Calibri')

wb.save(OUTPUT)
print(f"\n✓  Saved → {OUTPUT}")
print(f"   Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
