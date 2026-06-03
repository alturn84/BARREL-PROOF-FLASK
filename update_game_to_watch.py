#!/usr/bin/env python3
"""
update_game_to_watch.py — Barrel Proof Baseball
─────────────────────────────────────────────────
Reads schedule.json, standings.json, and mlb_pitching_stats.xlsx,
scores every game on today's slate, selects the Game to Watch,
and writes Site Data/game_to_watch.json.

Pipeline position:
    update_standings.py → update_schedule.py → update_game_to_watch.py

Cron:
    35 8 * * * /Library/Frameworks/Python.framework/Versions/3.14/bin/python3 \
        "update_game_to_watch.py"

Usage:
    python3 update_game_to_watch.py
"""

import json
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

print(f"SCRIPT STARTED: {datetime.now()}", flush=True)

VAULT        = Path(__file__).resolve().parent
SITE_DATA    = VAULT / "Site Data"
PITCHING_XLS = VAULT / "2026 Pitching Stats" / "mlb_pitching_stats.xlsx"
OUT_FILE     = SITE_DATA / "game_to_watch.json"

# ── Constants ─────────────────────────────────────────────────────────────────
LARGE_MARKET = {"NYY", "LAD", "BOS", "CHC", "ATL", "HOU", "SF", "NYM", "PHI"}

RIVALRIES = {
    frozenset({"NYY", "BOS"}), frozenset({"LAD", "SF"}),
    frozenset({"CHC", "STL"}), frozenset({"NYY", "NYM"}),
    frozenset({"ATL", "NYM"}), frozenset({"LAD", "SD"}),
    frozenset({"TEX", "HOU"}), frozenset({"ATH", "SF"}),
}

TEAM_NAMES = {
    "ARI": "Arizona Diamondbacks", "ATL": "Atlanta Braves",
    "BAL": "Baltimore Orioles",    "BOS": "Boston Red Sox",
    "CHC": "Chicago Cubs",         "CWS": "Chicago White Sox",
    "CIN": "Cincinnati Reds",      "CLE": "Cleveland Guardians",
    "COL": "Colorado Rockies",     "DET": "Detroit Tigers",
    "HOU": "Houston Astros",       "KC":  "Kansas City Royals",
    "LAA": "Los Angeles Angels",   "LAD": "Los Angeles Dodgers",
    "MIA": "Miami Marlins",        "MIL": "Milwaukee Brewers",
    "MIN": "Minnesota Twins",      "NYM": "New York Mets",
    "NYY": "New York Yankees",     "ATH": "Oakland Athletics",
    "PHI": "Philadelphia Phillies","PIT": "Pittsburgh Pirates",
    "SD":  "San Diego Padres",     "SF":  "San Francisco Giants",
    "SEA": "Seattle Mariners",     "STL": "St. Louis Cardinals",
    "TB":  "Tampa Bay Rays",       "TEX": "Texas Rangers",
    "TOR": "Toronto Blue Jays",    "WSH": "Washington Nationals",
}


# ── Helpers ───────────────────────────────────────────────────────────────────
def load_json(path):
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {}


def normalize(name):
    """Lowercase, strip accents, collapse whitespace for fuzzy name matching."""
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", name)
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(ascii_str.lower().split())


def load_era_lookup():
    """Returns {normalized_name: era_float} from the pitching XLSX."""
    if not openpyxl or not PITCHING_XLS.exists():
        print("  ⚠ openpyxl not available or XLSX missing — pitching scores skipped")
        return {}
    wb = openpyxl.load_workbook(PITCHING_XLS, read_only=True, data_only=True)
    ws = wb["MLB Pitching 2026"]
    headers = [cell.value for cell in next(ws.iter_rows())]
    name_idx = headers.index("Name")
    era_idx  = headers.index("ERA")
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        era  = row[era_idx]
        if name and era is not None:
            try:
                lookup[normalize(str(name))] = float(era)
            except (ValueError, TypeError):
                pass
    print(f"  Loaded ERA data for {len(lookup)} pitchers")
    return lookup


def get_era(era_lookup, pitcher_name):
    """Look up ERA by name with accent-insensitive matching."""
    if not pitcher_name or not era_lookup:
        return None
    return era_lookup.get(normalize(pitcher_name))


# ── Scoring ───────────────────────────────────────────────────────────────────
def score_game(game, standings, era_lookup):
    away = game.get("away_abbr", "")
    home = game.get("home_abbr", "")
    away_pitcher = game.get("away_prob", "")
    home_pitcher = game.get("home_prob", "")

    breakdown = {}
    total = 0

    def add(key, pts):
        nonlocal total
        breakdown[key] = pts
        total += pts

    # ── Pitching matchup ──────────────────────────────────────────────────────
    away_era = get_era(era_lookup, away_pitcher)
    home_era = get_era(era_lookup, home_pitcher)

    if away_era is not None and home_era is not None:
        combined = (away_era + home_era) / 2
        if combined <= 3.00:
            add("pitching_matchup", 40)
        elif combined <= 3.50:
            add("pitching_matchup", 30)
        elif combined <= 4.00:
            add("pitching_matchup", 20)
        else:
            add("pitching_matchup", 8)
    elif away_era is not None or home_era is not None:
        known = away_era if away_era is not None else home_era
        if known <= 3.00:
            add("pitching_matchup", 20)
        else:
            add("pitching_matchup", 8)

    # ── Standings context ─────────────────────────────────────────────────────
    away_std = standings.get(away, {})
    home_std = standings.get(home, {})
    away_gb  = away_std.get("gb", 99.0)
    home_gb  = home_std.get("gb", 99.0)
    away_div = away_std.get("division_id")
    home_div = home_std.get("division_id")
    same_div = away_div == home_div and away_div is not None

    if away_std.get("div_rank") == "1" and home_std.get("div_rank") == "1":
        add("standings_context", 35)
    elif (away_std.get("div_rank") == "1" and home_gb <= 2.0) or \
         (home_std.get("div_rank") == "1" and away_gb <= 2.0):
        add("standings_context", 28)
    elif away_gb <= 2.0 and home_gb <= 2.0:
        add("standings_context", 22)
    elif same_div:
        add("standings_context", 15)

    # ── Rivalry ───────────────────────────────────────────────────────────────
    if frozenset({away, home}) in RIVALRIES:
        add("rivalry", 20)

    # ── National relevance ────────────────────────────────────────────────────
    if home in LARGE_MARKET:
        add("national_relevance", 10)
    elif away in LARGE_MARKET:
        add("national_relevance", 5)

    # ── Weekend bonus ─────────────────────────────────────────────────────────
    dow = datetime.today().weekday()  # 0=Mon, 4=Fri, 5=Sat, 6=Sun
    if dow >= 4:
        add("weekend", 5)

    return total, breakdown


def build_standings_lookup():
    """
    Parse standings.json into {abbr: {div_rank, gb, division_id}} —
    mirrors the logic in update_game_of_day.py.
    """
    data = load_json(SITE_DATA / "standings.json")
    result = {}
    for league in data.get("leagues", []):
        for div in league.get("divisions", []):
            div_id = div.get("id", div.get("name", ""))
            for i, team in enumerate(div.get("teams", []), start=1):
                abbr = team.get("abbr", "")
                gb_raw = team.get("gb", "0")
                try:
                    gb = 0.0 if gb_raw in ("-", "", None) else float(gb_raw)
                except (ValueError, TypeError):
                    gb = 99.0
                result[abbr] = {
                    "div_rank":    str(i),
                    "gb":          gb,
                    "division_id": div_id,
                }
    return result


# ── Editorial copy ────────────────────────────────────────────────────────────
def build_copy(game, breakdown, away_era, home_era):
    away = game.get("away_abbr", "")
    home = game.get("home_abbr", "")
    away_full = TEAM_NAMES.get(away, away)
    home_full = TEAM_NAMES.get(home, home)
    away_pitcher = game.get("away_prob", "TBD")
    home_pitcher = game.get("home_prob", "TBD")

    is_rivalry   = "rivalry"          in breakdown
    is_standings = breakdown.get("standings_context", 0) >= 28
    has_aces     = breakdown.get("pitching_matchup", 0) >= 30

    # Headline
    if is_rivalry and has_aces:
        headline = f"{away_pitcher.split()[-1]} vs. {home_pitcher.split()[-1]}."
    elif is_rivalry:
        headline = f"{away_full.split()[-1]} at {home_full.split()[-1]}."
    elif has_aces:
        headline = f"{away_pitcher.split()[-1]} vs. {home_pitcher.split()[-1]}."
    elif is_standings:
        headline = f"Division stakes in {home_full.split()[-1]}."
    else:
        headline = f"{away_full} at {home_full}."

    # Subheadline
    parts = []
    if has_aces and away_era and home_era:
        parts.append(
            f"{away_pitcher} ({away_era:.2f} ERA) vs. "
            f"{home_pitcher} ({home_era:.2f} ERA)"
        )
    elif away_pitcher and home_pitcher:
        parts.append(f"{away_pitcher} vs. {home_pitcher}")
    if is_standings:
        parts.append("division implications")
    if is_rivalry:
        parts.append("rivalry game")
    subheadline = " · ".join(parts) if parts else f"{away_full} at {home_full}"

    # Reason
    reasons = []
    if has_aces:
        reasons.append(
            f"One of the better pitching matchups on the slate"
            + (f" — combined ERA of {((away_era or 0) + (home_era or 0)) / 2:.2f}" if away_era and home_era else "")
        )
    if is_rivalry:
        reasons.append("a natural rivalry adds stakes beyond the standings")
    if is_standings:
        reasons.append("both clubs are in the thick of the division race")
    if not reasons:
        reasons.append("best composite score on today's slate")
    reason = ". ".join(r.capitalize() for r in reasons) + "."

    return headline, subheadline, reason


# ── Main ──────────────────────────────────────────────────────────────────────
def run():
    schedule  = load_json(SITE_DATA / "schedule.json")
    era_lookup = load_era_lookup()
    standings  = build_standings_lookup()

    games = schedule.get("today", {}).get("games", [])
    # Filter to scheduled/preview games only — skip Final/Live
    upcoming = [g for g in games if g.get("status", "").lower()
                not in ("final", "live", "in progress")]

    if not upcoming:
        # Fall back to full slate if all games already started
        upcoming = games

    if not upcoming:
        print("  No games found in schedule.json")
        return

    print(f"  Scoring {len(upcoming)} game(s)...")

    scored = []
    for game in upcoming:
        pts, breakdown = score_game(game, standings, era_lookup)
        scored.append((pts, breakdown, game))

    # Sort: score desc, then rivalry, then standings, then game_pk asc
    scored.sort(key=lambda x: (
        -x[0],
        -x[1].get("rivalry", 0),
        -x[1].get("standings_context", 0),
        x[2].get("game_pk", 0),
    ))

    pts, breakdown, winner = scored[0]

    away = winner.get("away_abbr", "")
    home = winner.get("home_abbr", "")
    away_pitcher = winner.get("away_prob", "TBD")
    home_pitcher = winner.get("home_prob", "TBD")
    away_era = get_era(era_lookup, away_pitcher)
    home_era = get_era(era_lookup, home_pitcher)

    headline, subheadline, reason = build_copy(winner, breakdown, away_era, home_era)

    # Rail date from schedule for display
    rail_date = schedule.get("rail_date", "")

    output = {
        "date":           schedule.get("games_date", ""),
        "date_display":   rail_date,
        "updated":        datetime.now().strftime("%Y-%m-%d %H:%M"),
        "away":           away,
        "home":           home,
        "away_full":      TEAM_NAMES.get(away, away),
        "home_full":      TEAM_NAMES.get(home, home),
        "game_time":      winner.get("time", ""),
        "away_pitcher":   away_pitcher,
        "home_pitcher":   home_pitcher,
        "away_era":       f"{away_era:.2f}" if away_era is not None else "—",
        "home_era":       f"{home_era:.2f}" if home_era is not None else "—",
        "headline":       headline,
        "subheadline":    subheadline,
        "reason":         reason,
        "score":          pts,
        "breakdown":      breakdown,
    }

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(
        json.dumps(output, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(f"\n  ✓  Game to Watch: {away} @ {home}  ({pts} pts)")
    print(f"     {headline}")
    print(f"     {subheadline}")
    print(f"     Saved → {OUT_FILE}")


if __name__ == "__main__":
    run()
