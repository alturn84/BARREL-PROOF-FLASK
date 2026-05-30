#!/usr/bin/env python3
"""
MLB Batting Stats Updater — Barrel Proof Vault
───────────────────────────────────────────────
Pulls current-season batting stats from Baseball Reference via pybaseball
and writes/overwrites a formatted Excel file.

UPDATED: Preserves team abbreviations to disambiguate NYY/NYM, CHC/CHW, LAA/LAD

Setup (one time):
    pip install pybaseball openpyxl pandas

Usage:
    python update_batting_stats.py              # current season
    python update_batting_stats.py 2025         # specific season

Schedule with cron (runs at 9 AM daily):
    0 9 * * * /Library/Frameworks/Python.framework/Versions/3.14/bin/python3 "/Users/allanturner/Documents/BARREL PROOF/update_batting_stats.py"
"""

import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pybaseball import batting_stats_bref
from pybaseball import cache

# ── Config ────────────────────────────────────────────────────────────────────
VAULT   = Path("/Users/allanturner/Documents/BARREL PROOF")
OUTPUT  = VAULT / "2026 Batting Stats" / "mlb_batting_stats.xlsx"
MIN_PA  = 50

cache.enable()
cache.purge()

# Team abbreviation mapping (Baseball Reference → Full Team Name)
TEAM_MAP = {
    # AL East
    'BAL': 'Baltimore Orioles',
    'BOS': 'Boston Red Sox',
    'NYY': 'New York Yankees',
    'TBR': 'Tampa Bay Rays',
    'TOR': 'Toronto Blue Jays',
    # AL Central
    'CHW': 'Chicago White Sox',
    'CLE': 'Cleveland Guardians',
    'DET': 'Detroit Tigers',
    'KCR': 'Kansas City Royals',
    'MIN': 'Minnesota Twins',
    # AL West
    'HOU': 'Houston Astros',
    'LAA': 'Los Angeles Angels',
    'OAK': 'Oakland Athletics',
    'SEA': 'Seattle Mariners',
    'TEX': 'Texas Rangers',
    # NL East
    'ATL': 'Atlanta Braves',
    'MIA': 'Miami Marlins',
    'NYM': 'New York Mets',
    'PHI': 'Philadelphia Phillies',
    'WSN': 'Washington Nationals',
    # NL Central
    'CHC': 'Chicago Cubs',
    'CIN': 'Cincinnati Reds',
    'MIL': 'Milwaukee Brewers',
    'PIT': 'Pittsburgh Pirates',
    'STL': 'St. Louis Cardinals',
    # NL West
    'ARI': 'Arizona Diamondbacks',
    'COL': 'Colorado Rockies',
    'LAD': 'Los Angeles Dodgers',
    'SDP': 'San Diego Padres',
    'SFG': 'San Francisco Giants',
}

def fix_name(val):
    """Fix pybaseball literal \\xNN escape sequences and strip B-Ref symbols."""
    if not isinstance(val, str):
        return val
    try:
        val = bytes(val, 'utf-8').decode('unicode_escape').encode('latin-1').decode('utf-8')
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    val = unicodedata.normalize('NFC', val)
    val = val.replace('*', '').replace('#', '').strip()
    return val

def map_team(abbr):
    """Map Baseball Reference team abbreviation to full name."""
    if pd.isna(abbr) or abbr == '':
        return ''
    if abbr in ['TOT', '2TM', '3TM']:
        return 'Multiple Teams'
    abbr = str(abbr).strip()
    return TEAM_MAP.get(abbr, f"{abbr} (Unknown)")

# ── Fetch ─────────────────────────────────────────────────────────────────────
season = int(sys.argv[1]) if len(sys.argv) > 1 else datetime.today().year
print(f"Fetching {season} batting stats from Baseball Reference...")

df = batting_stats_bref(season)

keep = ['Name','Age','Tm','G','PA','AB','R','H','2B','3B','HR','RBI',
        'BB','IBB','SO','HBP','SH','SF','GDP','SB','CS','BA','OBP','SLG','OPS']

df = df[[c for c in keep if c in df.columns]].copy()
df = df.rename(columns={'GDP': 'GIDP'})
df['Name'] = df['Name'].apply(fix_name)
df['Team'] = df['Tm'].apply(map_team)
df = df.drop(columns=['Tm'])

df = df[df['PA'] >= MIN_PA].sort_values('PA', ascending=False).reset_index(drop=True)
df.insert(0, 'Rk', range(1, len(df) + 1))

print(f"  {len(df)} players with {MIN_PA}+ PA found.")

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"MLB Batting {season}"

header_fill  = PatternFill('solid', start_color='1F4E79')
header_font  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
alt_fill     = PatternFill('solid', start_color='D6E4F0')
white_fill   = PatternFill('solid', start_color='FFFFFF')
body_font    = Font(name='Calibri', size=10)
center_align = Alignment(horizontal='center', vertical='center')
left_align   = Alignment(horizontal='left',   vertical='center')
thin         = Side(style='thin', color='AAAAAA')
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

headers  = list(df.columns)
avg_cols = {'BA','OBP','SLG','OPS'}

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = border

ws.row_dimensions[1].height = 22

for row_idx, (_, row) in enumerate(df.iterrows(), 2):
    fill = alt_fill if row_idx % 2 == 0 else white_fill
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val  = row[col_name]
        if hasattr(val, 'item'):
            val = val.item()
        cell.value     = val
        cell.font      = body_font
        cell.fill      = fill
        cell.border    = border
        cell.alignment = left_align if col_name in ['Name', 'Team'] else center_align
        if col_name in avg_cols and isinstance(val, float):
            cell.number_format = '0.000'

widths = {
    'Rk':5,'Name':24,'Age':5,'Team':22,'G':5,'PA':5,'AB':5,
    'R':5,'H':5,'2B':5,'3B':5,'HR':5,'RBI':5,'BB':5,'IBB':5,
    'SO':5,'HBP':5,'SH':5,'SF':5,'GIDP':6,'SB':5,'CS':5,
    'BA':7,'OBP':7,'SLG':7,'OPS':7
}
for col_idx, col_name in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 7)

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

meta = wb.create_sheet("Info")
meta['A1'] = "Last updated"
meta['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M")
meta['A2'] = "Season"
meta['B2'] = season
meta['A3'] = "Min PA filter"
meta['B3'] = MIN_PA
meta['A4'] = "Source"
meta['B4'] = "Baseball Reference via pybaseball"
meta['A5'] = "Teams"
meta['B5'] = "Full team names (disambiguated)"
for c in ['A1','A2','A3','A4','A5']:
    meta[c].font = Font(bold=True, name='Calibri')

wb.save(OUTPUT)
print(f"\n✓  Saved → {OUTPUT}")
print(f"   Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
